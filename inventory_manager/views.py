# inventory_manager/views.py
import os
from django.shortcuts import render, redirect, get_object_or_404
from django.core.serializers.json import DjangoJSONEncoder
import json
from django.contrib import messages
from django.db.models import Q, Sum, Count
import openpyxl
from openpyxl.utils import get_column_letter
from .models import SKU, InventoryTransaction, ChannelListing, SyncLog, Bag, BagItem, Rack
from .forms import ManualAdjustForm, OrderFileUploadForm, SKUCreateForm, BagForm, BagItemFormSet, BagItemForm
from django.forms import formset_factory
import pandas as pd
from django.http import HttpResponse
from datetime import datetime
from django.db import models
from django.core.paginator import Paginator
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.utils.timezone import now
from django.utils.http import urlencode
from io import BytesIO
from django.core.cache import cache
from django.views.decorators.http import require_POST
import threading, time
from django.contrib.auth import login
from django.contrib.auth.forms import UserCreationForm
from inventory_manager.flipkart_api import sync_inventory_to_flipkart, get_flipkart_token
from django.db import IntegrityError,transaction
from django.contrib.auth.decorators import login_required


def register(request):
    if request.method == "POST":
        form = UserCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            return redirect('dashboard')
    else:
        form = UserCreationForm()
    return render(request, 'inventory_manager/register.html', {'form': form})

@login_required
def sync_all_inventory(request):
    skus = SKU.objects.filter(user=request.user)
    success = 0
    failed = 0

    for sku in skus:
        try:
            sync_inventory_to_flipkart(sku)
            success += 1
        except Exception as e:
            print(f"❌ Failed to sync {sku.sku_code}: {str(e)}")
            failed += 1

    messages.success(request, f"Flipkart sync complete: {success} succeeded, {failed} failed.")
    return redirect('dashboard')

# Track sync progress globally
sync_progress = {"progress": 0}

@login_required
@csrf_exempt
def start_inventory_sync(request):
    if request.method == "POST":
        user = request.user
        selected_product = request.POST.get("product_name")

        def perform_sync():
            skus = SKU.objects.filter(product_name=selected_product) if selected_product and selected_product != "All Products" else SKU.objects.filter(user=request.user)

            total = skus.count()
            updated = 0
            sync_progress["progress"] = 0

            try:
                token = get_flipkart_token()
            except Exception as e:
                sync_progress["progress"] = 100
                return

            for i, sku in enumerate(skus, start=1):
                try:
                    if sync_inventory_to_flipkart(sku, token):
                        updated += 1
                        print(updated)
                except Exception as e:
                    print(f"Error syncing {sku.sku_code}: {e}")
                sync_progress["progress"] = int((i / total) * 100)
            print(updated)
            SyncLog.objects.create(
                user=user,
                synced_items=updated,
                status="completed",
                product_name=selected_product or "All Products"
            )

        threading.Thread(target=perform_sync).start()
        return JsonResponse({"message": "Sync started"})

    return JsonResponse({"error": "Invalid method"}, status=400)

@login_required
def get_sync_progress(request):
    return JsonResponse({"progress": sync_progress.get("progress", 0)})

@login_required
def sync_logs_view(request):
    query = request.GET.get("q", "").strip()
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")
    per_page = int(request.GET.get("per_page", 10))

    logs = SyncLog.objects.filter(user=request.user)

    if query:
        logs = logs.filter(product_name__icontains=query)
    if start_date:
        try:
            start = datetime.strptime(start_date, "%Y-%m-%d").date()
            logs = logs.filter(timestamp__date__gte=start)
        except ValueError:
            pass  # ignore invalid date
    if end_date:
        try:
            end = datetime.strptime(end_date, "%Y-%m-%d").date()
            logs = logs.filter(timestamp__date__lte=end)
        except ValueError:
            pass  # ignore invalid date

    logs = logs.order_by("-timestamp")

    # ✅ Export
    if "export" in request.GET:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sync Logs"
        ws.append(["Timestamp", "Product Name", "Synced Items", "Status"])

        for log in logs:
            ws.append([
                log.timestamp.strftime("%Y-%m-%d %H:%M"),
                log.product_name,
                log.synced_items,
                log.status
            ])

        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = "attachment; filename=sync_logs.xlsx"
        wb.save(response)
        return response

    # ✅ Pagination
    paginator = Paginator(logs, per_page)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    context = {
        "logs": page_obj,
        "query": query,
        "start_date": start_date,
        "end_date": end_date,
        "page_obj": page_obj,
        "per_page": per_page,
        "request": request
    }
    return render(request, "inventory_manager/sync_logs.html", context)

@login_required
def export_sync_logs(request):
    query = request.GET.get("q", "").strip()

    logs = SyncLog.objects.filter(user=request.user)
    if query:
        logs = logs.filter(product_name__icontains=query)

    logs = logs.order_by("-timestamp")

    df = pd.DataFrame.from_records(
        logs.values("timestamp", "product_name", "synced_items", "status")
    )
    df.rename(columns={
        "timestamp": "Timestamp",
        "product_name": "Product Name",
        "synced_items": "Synced Items",
        "status": "Status"
    }, inplace=True)

    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="xlsxwriter") as writer:
        df.to_excel(writer, index=False, sheet_name="Sync Logs")

    buffer.seek(0)
    response = HttpResponse(
        buffer,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=sync_logs.xlsx"
    return response

@login_required                        
def dashboard(request):
    skus = SKU.objects.filter(user=request.user)
    total_skus = skus.count()
    total_stock = sum(s.stock_quantity for s in skus)
    low_stock = skus.filter(stock_quantity__lt=10, stock_quantity__gt=0).count()
    out_of_stock = skus.filter(stock_quantity=0).count()
    high_stock = skus.filter(stock_quantity__gte=10).count()

    top_skus = skus.order_by('-stock_quantity')[:10]
    top_skus_labels = [sku.sku_code for sku in top_skus]
    top_skus_data = [sku.stock_quantity for sku in top_skus]

    categories = sorted(set(sku.product_name.split()[0] for sku in skus if sku.product_name))
    sync_logs = SyncLog.objects.filter(user=request.user).order_by('-timestamp')[:5]

    product_names = (SKU.objects.filter(user=request.user).values_list("product_name", flat=True).distinct().order_by("product_name"))


    return render(request, "inventory_manager/dashboard.html", {
        "total_skus": total_skus,
        "total_stock": total_stock,
        "low_stock": low_stock,
        "out_of_stock": out_of_stock,
        "high_stock": high_stock,
        "top_skus_labels": top_skus_labels,
        "top_skus_data": top_skus_data,
        "categories": categories,
        "sync_logs": sync_logs,
        "product_names": product_names,
    })

@login_required
def master_skus(request):
    # GET parameters
    query = request.GET.get("q", "").strip()
    sort = request.GET.get("sort", "sku_code")
    dir = request.GET.get("dir", "asc")
    per_page = int(request.GET.get("per_page", 25))
    page_number = request.GET.get("page", 1)

    # Base queryset
    skus = SKU.objects.filter(user=request.user)

    # Search filter
    if query:
        skus = skus.filter(
            Q(sku_code__icontains=query) |
            Q(product_name__icontains=query)
        )

    # Validate sort inputs
    allowed_fields = ["sku_code", "product_name", "stock_quantity"]
    if sort not in allowed_fields:
        sort = "sku_code"
    if dir not in ["asc", "desc"]:
        dir = "asc"

    order_by = sort if dir == "asc" else f"-{sort}"
    skus = skus.filter(user=request.user).order_by(order_by)

    # ✅ Excel export (if requested)
    if request.GET.get("export") == "1":
        export_qs = skus  # already filtered + sorted
        df = pd.DataFrame(export_qs.values("sku_code", "product_name", "stock_quantity", "updated_at"))

        # Remove timezone info from datetime column
        if 'updated_at' in df.columns:
            df['updated_at'] = pd.to_datetime(df['updated_at']).dt.tz_localize(None)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="Master_SKUs.xlsx"'
        df.to_excel(response, index=False)
        return response

    # Pagination
    paginator = Paginator(skus, per_page)
    page_obj = paginator.get_page(page_number)

    return render(request, "inventory_manager/master_skus.html", {
        "skus": page_obj,
        "page_obj": page_obj,
        "query": query,
        "sort": sort,
        "dir": dir,
        "per_page": per_page,
    })

@login_required
def upload_channel_listing(request):
    if request.method == 'POST':
        file = request.FILES.get('file')
        if not file:
            return render(request, 'inventory_manager/upload_channel_listing.html', {'error': 'No file selected.'})

        try:
            df = pd.read_excel(file, skiprows=[1])  # ✅ Skip second row (index 1)
        except Exception as e:
            return render(request, 'inventory_manager/upload_channel_listing.html', {'error': f"Error reading Excel: {e}"})

        required_columns = ['Seller SKU Id', 'Flipkart Serial Number']
        if not all(col in df.columns for col in required_columns):
            return render(request, 'inventory_manager/upload_channel_listing.html', {
                'error': 'Missing required columns: Seller SKU Id and Flipkart Serial Number.'
            })

        for _, row in df.iterrows():
            sku = str(row['Seller SKU Id']).strip()
            fsn = str(row['Flipkart Serial Number']).strip()

            if sku and fsn:
                ChannelListing.objects.get_or_create(user=request.user, channel_sku=sku, defaults={"fsn": fsn})

        return redirect('view_channel_listings')

    return render(request, 'inventory_manager/upload_channel_listing.html')

@login_required
def bulk_map_channel_skus(request):
    if request.method == "POST":
        file = request.FILES.get("file")
        if not file:
            messages.error(request, "No file uploaded.")
            return redirect("view_channel_listings")

        try:
            df = pd.read_excel(file)
        except Exception as e:
            messages.error(request, f"Failed to read file: {e}")
            return redirect("view_channel_listings")

        required_columns = ["Seller SKU Id", "Master SKU Code"]
        if not all(col in df.columns for col in required_columns):
            messages.error(request, "Missing required columns: 'Seller SKU Id' and 'Master SKU Code'.")
            return redirect("view_channel_listings")

        updated = 0
        failed_rows = []

        for _, row in df.iterrows():
            channel_sku = str(row["Seller SKU Id"]).strip()
            master_code = str(row["Master SKU Code"]).strip()

            # Check for ChannelListing
            listing = ChannelListing.objects.filter(user=request.user, channel_sku=channel_sku).first()
            if not listing:
                failed_rows.append({
                    "Seller SKU Id": channel_sku,
                    "Flipkart Serial Number": "",
                    "Master SKU Code": master_code,
                    "Error Message": "Channel SKU not found"
                })
                continue

            # Check for Master SKU
            master_sku = SKU.objects.filter(sku_code=master_code).first()
            if not master_sku:
                failed_rows.append({
                    "Seller SKU Id": channel_sku,
                    "Flipkart Serial Number": listing.fsn,
                    "Master SKU Code": master_code,
                    "Error Message": "Master SKU not found"
                })
                continue

            # All good: Map
            listing.master_sku = master_sku
            listing.save()
            updated += 1

        if failed_rows:
            request.session["channel_mapping_failures"] = failed_rows
            messages.warning(request, f"✅ {updated} mapped. ❌ {len(failed_rows)} failed. You can download the failed rows below.")
        else:
            messages.success(request, f"✅ {updated} mappings updated successfully.")

        return redirect("view_channel_listings")

    return redirect("view_channel_listings")

@login_required
def download_failed_channel_mappings(request):
    failed_rows = request.session.get("channel_mapping_failures")
    if not failed_rows:
        messages.error(request, "No failed rows found.")
        return redirect("view_channel_listings")

    # Clear session after retrieving
    del request.session["channel_mapping_failures"]

    df_failed = pd.DataFrame(failed_rows)
    output = BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df_failed.to_excel(writer, index=False, sheet_name='Failed_Mappings')

    output.seek(0)
    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="Failed_Channel_Mappings.xlsx"'
    return response

@login_required
def download_channel_mapping_template(request):

    # Fetch all Channel Listings
    all_listings = ChannelListing.objects.filter(user=request.user).select_related("master_sku")

    # Split into unmapped and mapped
    unmapped_data = []
    mapped_data = []

    for cl in all_listings:
        row = {
            "Seller SKU Id": cl.channel_sku,
            "Flipkart Serial Number": cl.fsn,
            "Master SKU Code": cl.master_sku.sku_code if cl.master_sku else ""
        }
        if cl.master_sku:
            mapped_data.append(row)
        else:
            unmapped_data.append(row)

    # Convert to DataFrames
    df_unmapped = pd.DataFrame(unmapped_data)
    df_mapped = pd.DataFrame(mapped_data)

    # Save to Excel in memory
    output = BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df_unmapped.to_excel(writer, index=False, sheet_name="Unmapped_SKUs")
        df_mapped.to_excel(writer, index=False, sheet_name="Mapped_SKUs")

    output.seek(0)
    response = HttpResponse(
        output,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="Channel_SKU_Mapping_Template.xlsx"'
    return response

@login_required
def view_channel_listings(request):
    query = request.GET.get("q", "")
    per_page = int(request.GET.get("per_page", 25))
    sort = request.GET.get("sort", "channel_sku")
    direction = request.GET.get("dir", "asc")
    filter_val = request.GET.get("filter", "all")
    page_number = request.GET.get("page")

    listings = ChannelListing.objects.filter(user=request.user).select_related("master_sku")

    # 🔍 Apply filter
    if filter_val == "mapped":
        listings = listings.exclude(master_sku__isnull=True)
    elif filter_val == "unmapped":
        listings = listings.filter(master_sku__isnull=True)

    # 🔍 Apply search
    if query:
        listings = listings.filter(
            Q(channel_sku__icontains=query) |
            Q(fsn__icontains=query) |
            Q(master_sku__sku_code__icontains=query) |
            Q(master_sku__product_name__icontains=query)
        )

    # ↕️ Apply sorting
    if direction == "desc":
        sort = f"-{sort}"
    listings = listings.order_by(sort)

    # 📄 Pagination
    paginator = Paginator(listings, per_page)
    page_obj = paginator.get_page(page_number)

    # ✅ POST Actions: Save, Bulk Delete, Reset All
    if request.method == "POST":
        updated = 0
        selected_ids = request.POST.getlist("selected_ids")

        if "bulk_delete" in request.POST:
            ChannelListing.objects.filter(user=request.user, id__in=selected_ids).delete()
            messages.success(request, f"{len(selected_ids)} listings deleted successfully.")
        elif "reset_all" in request.POST:
            for listing in page_obj.object_list:
                listing.master_sku = None
                listing.save()
            messages.success(request, "Mappings reset for this page.")
        else:
            for listing in page_obj.object_list:
                master_sku_id = request.POST.get(f"mapping_{listing.id}")
                if master_sku_id:
                    try:
                        listing.master_sku = SKU.objects.get(id=master_sku_id)
                        listing.save()
                        updated += 1
                    except SKU.DoesNotExist:
                        pass
            messages.success(request, f"{updated} mappings saved successfully.")

        # 🔁 Redirect to same page with preserved query
        query_dict = {
            "q": query,
            "per_page": per_page,
            "sort": sort.lstrip("-"),
            "dir": direction,
            "filter": filter_val,
            "page": page_number,
        }
        return redirect(f"{request.path}")

    return render(request, "inventory_manager/map_channel_listings.html", {
        "listings": page_obj.object_list,
        "skus": SKU.objects.filter(user=request.user),
        "query": query,
        "sort": sort.lstrip("-"),
        "dir": direction,
        "filter": filter_val,
        "page_obj": page_obj,
        "per_page": per_page,
    })

@login_required
@csrf_exempt
def update_stock(request):
    if request.method == "POST":
        sku_id = request.POST.get("sku_id")
        new_qty = request.POST.get("quantity")
        notes = request.POST.get("notes", "Inline update")

        try:
            sku = SKU.objects.filter(user=request.user).get(id=sku_id)
            new_qty = int(new_qty)
            delta = new_qty - sku.stock_quantity
            if delta != 0:
                # Log transaction
                InventoryTransaction.objects.create(
                    user=request.user,
                    sku=sku,
                    quantity_change=delta,
                    source="manual",
                    notes=notes,
                    timestamp=now()
                )
            sku.stock_quantity = new_qty
            sku.save()
            return JsonResponse({"success": True, "quantity": sku.stock_quantity})
        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)})

    return JsonResponse({"success": False, "error": "POST required"})

@login_required
def manual_adjust(request):
    if request.method == 'POST':
        form = ManualAdjustForm(request.POST)
        reset = request.POST.get('reset') == 'on'
        if form.is_valid():
            sku = form.cleaned_data['sku']
            quantity = form.cleaned_data['quantity_change']
            notes = form.cleaned_data['notes']

            if reset:
                difference = quantity - sku.stock_quantity
                sku.stock_quantity = quantity
                source = 'reset'
                notes = f"Reset to {quantity}. {notes or ''}".strip()
            else:
                sku.stock_quantity += quantity
                difference = quantity
                source = 'manual'

            sku.save()
            InventoryTransaction.objects.create(
                user=request.user,
                sku=sku,
                quantity_change=difference,
                source=source,
                notes=notes
            )
            messages.success(request, f'Stock updated for {sku.sku_code}.')
            return redirect('manual_adjust')
    else:
        form = ManualAdjustForm()

    # Prepare data for stock display
    stock_data = {str(sku.pk): sku.stock_quantity for sku in SKU.objects.filter(user=request.user)}
    return render(request, 'inventory_manager/manual_adjust.html', {
        'form': form,
        'stock_data': json.dumps(stock_data, cls=DjangoJSONEncoder)
    })

@login_required
def upload_excel(request):
    if request.method == 'POST':
        file = request.FILES.get('excel_file')
        reset_stock = 'reset_stock' in request.POST  # ✅ checkbox value

        if not file:
            return render(request, 'inventory_manager/upload_excel.html', {'error': 'No file uploaded.'})

        try:
            df = pd.read_excel(file)
        except Exception as e:
            return render(request, 'inventory_manager/upload_excel.html', {'error': f'Error reading file: {e}'})

        required_columns = ['SKU Code', 'Quantity']
        for col in required_columns:
            if col not in df.columns:
                return render(request, 'inventory_manager/upload_excel.html', {'error': f'Missing required column: {col}'})

        for _, row in df.iterrows():
            sku_code = str(row['SKU Code']).strip()
            qty = int(row['Quantity'])

            try:
                sku = SKU.objects.get(user=request.user,sku_code=sku_code)
                if reset_stock:
                    # ✅ Reset mode: overwrite quantity
                    change = qty - sku.stock_quantity
                    sku.stock_quantity = qty
                    sku.save()
                    if change != 0:
                        InventoryTransaction.objects.create(
                            user=request.user,
                            sku=sku,
                            quantity_change=change,
                            source='excel',
                            notes='Reset via stock Excel upload'
                        )
                else:
                    # ➕ Add mode: increase quantity
                    sku.stock_quantity += qty
                    sku.save()
                    InventoryTransaction.objects.create(
                        user=request.user,
                        sku=sku,
                        quantity_change=qty,
                        source='excel',
                        notes='Added via stock Excel upload'
                    )
            except SKU.DoesNotExist:
                continue  # Skip unknown SKUs or optionally log them

        messages.success(request, "✅ Stock updated successfully.")
        return redirect('dashboard')

    return render(request, 'inventory_manager/upload_excel.html')

@login_required
def upload_order_file(request):
    if request.method == 'POST':
        form = OrderFileUploadForm(request.POST, request.FILES)
        if form.is_valid():
            file_name = request.FILES['file']
            extension = file_name.name.lower()
            if extension.endswith('.csv'):
                df = pd.read_csv(request.FILES['file'])
            elif extension.endswith('.xls') or extension.endswith('.xlsx'):
                df = pd.read_excel(request.FILES['file'])
            else:
                raise ValueError("Unsupported file type")
            
            for _, row in df.iterrows():
                raw_sku = str(row.get('SKU')).strip()
                qty = int(row.get('Quantity', 0))

                # Try to match directly to Master SKU
                sku = SKU.objects.filter(user=request.user,sku_code=raw_sku).first()

                # If not found, look up in ChannelListing and get master SKU
                if not sku:
                    channel = ChannelListing.objects.filter(user=request.user, channel_sku=raw_sku).first()
                    if channel and channel.master_sku:
                        sku = channel.master_sku

                # If resolved, deduct stock and log transaction
                if sku:
                    sku.stock_quantity -= qty
                    sku.save()
                    InventoryTransaction.objects.create(
                        user=request.user,
                        sku=sku,
                        quantity_change=-qty,
                        source='flipkart_order',
                        notes=f'Deducted via file for: {raw_sku}'
                    )
            messages.success(request, "Orders deducted successfully.")
            return redirect('dashboard')
    else:
        form = OrderFileUploadForm()
    return render(request, 'inventory_manager/upload_orders.html', {'form': form})

@login_required
def add_sku(request):
    if request.method == 'POST':
        form = SKUCreateForm(request.POST)
        if form.is_valid():
            sku_code = form.cleaned_data['sku_code']
            product_name = form.cleaned_data['product_name']
            stock_quantity = form.cleaned_data['stock_quantity']

            if SKU.objects.filter(user=request.user, sku_code=sku_code).exists():
                form.add_error('sku_code', 'SKU already exists.')
            else:
                SKU.objects.create(
                    user=request.user,
                    sku_code=sku_code,
                    product_name=product_name,
                    stock_quantity=stock_quantity
                )
                return redirect('dashboard')
    else:
        form = SKUCreateForm()
    return render(request, 'inventory_manager/add_sku.html', {'form': form})

@login_required
def upload_sku_excel(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')

        if not excel_file:
            return render(request, 'inventory_manager/upload_sku_excel.html', {
                'error': 'No file uploaded.'
            })

        try:
            df = pd.read_excel(excel_file)
        except Exception as e:
            return render(request, 'inventory_manager/upload_sku_excel.html', {
                'error': f'Error reading file: {e}'
            })

        required_columns = ['SKU Code', 'Product Name', 'Stock']
        for col in required_columns:
            if col not in df.columns:
                return render(request, 'inventory_manager/upload_sku_excel.html', {
                    'error': f'Missing required column: {col}'
                })

        for _, row in df.iterrows():
            SKU.objects.update_or_create(
                user=request.user,
                sku_code=str(row['SKU Code']).strip(),
                defaults={
                    'product_name': str(row['Product Name']).strip(),
                    'stock_quantity': int(row['Stock'])
                }
            )
        messages.success(request, 'Uploaded Successfully')
        return redirect('dashboard')

    return render(request, 'inventory_manager/upload_sku_excel.html')

@login_required
def delete_channel_listing(request, pk):
    listing = get_object_or_404(ChannelListing, pk=pk, user=request.user)
    if request.method == "POST":
        listing.delete()
        messages.success(request, "Listing deleted.")
    return redirect("view_channel_listings")

@login_required
def view_mappings(request):
    query = request.GET.get("q", "").strip()
    sort = request.GET.get("sort", "channel_sku")
    dir = request.GET.get("dir", "asc")
    per_page = int(request.GET.get("per_page", 25))
    page_number = request.GET.get("page", 1)

    mappings = ChannelListing.objects.filter(user=request.user).select_related("master_sku")

    if query:
        mappings = mappings.filter(
            Q(channel_sku__icontains=query) |
            Q(fsn__icontains=query) |
            Q(master_sku__sku_code__icontains=query) |
            Q(master_sku__product_name__icontains=query)
        )

    # Sorting
    valid_fields = {
        "channel_sku": "channel_sku",
        "fsn": "fsn",
        "master_sku__sku_code": "master_sku__sku_code",
        "master_sku__product_name": "master_sku__product_name"
    }

    order = valid_fields.get(sort, "channel_sku")
    if dir == "desc":
        order = f"-{order}"
    mappings = mappings.order_by(order)

    # Export to Excel
    if request.GET.get("export") == "1":
        df = pd.DataFrame(list(mappings.values(
            "channel_sku", "fsn", "master_sku__sku_code", "master_sku__product_name"
        )))
        df.columns = ["Flipkart SKU", "FSN", "Master SKU Code", "Product Name"]

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="Channel_SKU_Mappings.xlsx"'
        df.to_excel(response, index=False)
        return response

    # Pagination
    paginator = Paginator(mappings, per_page)
    page_obj = paginator.get_page(page_number)

    return render(request, "inventory_manager/view_mappings.html", {
        "mappings": page_obj,
        "page_obj": page_obj,
        "query": query,
        "sort": sort,
        "dir": dir,
        "per_page": per_page,
        "skus": SKU.objects.filter(user=request.user)
    })

@login_required
def inventory_history(request):
    query = request.GET.get("q", "").strip()
    start_date = request.GET.get("start")
    end_date = request.GET.get("end")
    per_page = int(request.GET.get("per_page", 25))
    page_number = request.GET.get("page", 1)

    txns = InventoryTransaction.objects.filter(user=request.user).select_related("sku").order_by("-timestamp")

    if query:
        txns = txns.filter(
            Q(sku__sku_code__icontains=query) |
            Q(sku__product_name__icontains=query) |
            Q(source__icontains=query) |
            Q(notes__icontains=query)
        )

    if start_date:
        txns = txns.filter(timestamp__date__gte=start_date)
    if end_date:
        txns = txns.filter(timestamp__date__lte=end_date)

    # ✅ Export to Excel
    if request.GET.get("export") == "1":
        data = []
        for t in txns:
            data.append({
                "Date": t.timestamp.astimezone().strftime("%Y-%m-%d %I:%M %p"),
                "SKU Code": t.sku.sku_code,
                "Product Name": t.sku.product_name,
                "Quantity Change": t.quantity_change,
                "Source": t.get_source_display(),
                "Notes": t.notes or "",
            })

        df = pd.DataFrame(data)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="Inventory_History.xlsx"'
        df.to_excel(response, index=False)
        return response

    # Pagination
    paginator = Paginator(txns, per_page)
    page_obj = paginator.get_page(page_number)

    return render(request, "inventory_manager/inventory_history.html", {
        "page_obj": page_obj,
        "query": query,
        "start": start_date,
        "end": end_date,
        "per_page": per_page,
    })

@login_required
def low_stock_alert(request):
    threshold = int(request.GET.get("threshold", 5))
    per_page = int(request.GET.get("per_page", 25))
    page_number = request.GET.get("page", 1)

    sort = request.GET.get("sort", "stock_quantity")
    dir = request.GET.get("dir", "asc")

    # Build base queryset
    queryset = SKU.objects.filter(user=request.user,stock_quantity__lte=threshold)

    # Determine sort direction
    if sort not in ["sku_code", "product_name", "stock_quantity"]:
        sort = "stock_quantity"
    if dir not in ["asc", "desc"]:
        dir = "asc"

    order_by_field = sort if dir == "asc" else f"-{sort}"
    queryset = queryset.order_by(order_by_field)

    # Export to Excel
    if request.GET.get("export") == "1":
        df = pd.DataFrame(queryset.values("sku_code", "product_name", "stock_quantity"))
        df["Threshold"] = threshold

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="Low_Stock_SKUs.xlsx"'
        df.to_excel(response, index=False)
        return response

    # Paginate
    paginator = Paginator(queryset, per_page)
    page_obj = paginator.get_page(page_number)

    return render(request, "inventory_manager/low_stock_alert.html", {
        "page_obj": page_obj,
        "threshold": threshold,
        "per_page": per_page,
        "sort": sort,
        "dir": dir,
    })

@login_required
def add_bag(request):
    BagItemFormSet = formset_factory(BagItemForm, extra=0, can_delete=True)

    if request.method == 'POST':
        bag_form = BagForm(request.POST)
        formset = BagItemFormSet(request.POST)

        if bag_form.is_valid() and formset.is_valid():
            reflect = 'reflect_in_rack' in request.POST
            bag_number = bag_form.cleaned_data['bag_number']
            seen_skus = set()
            duplicate_found = False

            # 🔍 Check rack quantity before saving
            if reflect:
                for form in formset:
                    if not form.cleaned_data.get("sku") or not form.cleaned_data.get("quantity"):
                        continue
                    sku = form.cleaned_data["sku"]
                    qty = form.cleaned_data["quantity"]
                    rack_entry = Rack.objects.filter(user=request.user, sku=sku).first()
                    if not rack_entry or rack_entry.quantity < qty:
                        messages.error(request, f"Not enough rack quantity for {sku.sku_code}.")
                        return redirect("add_bag")

            try:
                bag = Bag.objects.create(user=request.user, bag_number=bag_number, reflect_in_rack=reflect)

                for form in formset:
                    if not form.cleaned_data.get("sku") or not form.cleaned_data.get("quantity"):
                        continue

                    sku = form.cleaned_data["sku"]
                    qty = form.cleaned_data["quantity"]

                    if sku.id in seen_skus:
                        duplicate_found = True
                        continue
                    seen_skus.add(sku.id)

                    # 💾 Save item
                    BagItem.objects.create(
                        bag=bag,
                        sku=sku,
                        quantity=qty,
                        product_name=sku.product_name
                    )

                    if reflect:
                        # 🟡 Deduct from rack
                        rack_entry = Rack.objects.get(user=request.user,sku=sku)
                        rack_entry.quantity -= qty
                        rack_entry.save()

                        InventoryTransaction.objects.create(
                            user=request.user,
                            sku=sku,
                            quantity_change=0,
                            source='bag_add',
                            notes=f'Moved {qty} from Rack to bag {bag.bag_number}'
                        )
                    else:
                        # 🟢 Add to master stock
                        sku.stock_quantity += qty
                        sku.save()

                        InventoryTransaction.objects.create(
                            user=request.user,
                            sku=sku,
                            quantity_change=qty,
                            source='bag_add',
                            notes=f'Added to bag {bag.bag_number}'
                        )

                if duplicate_found:
                    messages.warning(request, "Duplicate SKUs were skipped while saving the bag.")

                messages.success(request, "Bag saved and inventory updated.")
                return redirect("view_bags")

            except IntegrityError:
                messages.error(request, f"Bag with number '{bag_number}' already exists.")
        else:
            messages.error(request, "There were errors in your form.")

    else:
        bag_form = BagForm()
        formset = BagItemFormSet()

    return render(request, "inventory_manager/add_bag.html", {
        "bag_form": bag_form,
        "formset": formset
    })

@login_required
def ajax_sku_search(request):
    term = request.GET.get('term', '')
    qs = SKU.objects.filter(user=request.user,sku_code__icontains=term)[:10]
    results = [{"id": sku.id, "text": f"{sku.sku_code} - {sku.product_name}"} for sku in qs]
    return JsonResponse({"results": results})

@login_required
def view_bags(request):
    bags = Bag.objects.prefetch_related('items', 'items__sku').order_by('-created_at').filter(user=request.user)

    # Filters
    date_from = request.GET.get("from")
    date_to = request.GET.get("to")
    export = request.GET.get("export")
    query = request.GET.get("q", "").strip()

    if query:
        bags = bags.filter(
            Q(bag_number__icontains=query) |
            Q(items__sku__sku_code__icontains=query)
        ).distinct()
    if date_from:
        bags = bags.filter(created_at__date__gte=date_from)
    if date_to:
        bags = bags.filter(created_at__date__lte=date_to)

    # ✅ Export filtered bags
    if export:
        rows = []
        for bag in bags:
            for item in bag.items.all():
                rows.append({
                    "Bag Number": bag.bag_number,
                    "SKU": item.sku.sku_code,
                    "Product Name": item.product_name,
                    "Quantity": item.quantity,
                    "Date": bag.created_at.strftime("%Y-%m-%d %H:%M"),
                })
        df = pd.DataFrame(rows)
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=bags_export.xlsx'
        df.to_excel(response, index=False)
        return response

    # ✅ Add Pagination
    per_page = int(request.GET.get("per_page", 10))
    page = request.GET.get("page", 1)
    paginator = Paginator(bags, per_page)
    page_obj = paginator.get_page(page)

    return render(request, "inventory_manager/view_bags.html", {
        "bags": page_obj,
        "query": query,
        "date_from": date_from,
        "date_to": date_to,
        "per_page": per_page,
        "page_obj": page_obj,
    })

@login_required
def edit_bag(request, pk):
    bag = get_object_or_404(Bag, pk=pk, user=request.user)
    old_items = {item.sku_id: item.quantity for item in bag.items.all()}

    if request.method == 'POST':
        bag_form = BagForm(request.POST)
        formset = BagItemFormSet(request.POST)

        if bag_form.is_valid() and formset.is_valid():
            reflect = 'reflect_in_rack' in request.POST
            bag.bag_number = bag_form.cleaned_data['bag_number']
            bag.reflect_in_rack = reflect
            bag.save()

            existing_items = {item.sku.id: item for item in bag.items.all()}
            updated_skus = []

            for form in formset:
                if not form.cleaned_data:
                    continue

                sku = form.cleaned_data['sku']
                quantity = form.cleaned_data['quantity']
                updated_skus.append(sku.id)

                if sku.id in existing_items:
                    # ✅ Update existing item
                    existing_item = existing_items.pop(sku.id)
                    old_qty = existing_item.quantity
                    difference = quantity - old_qty

                    if reflect:
                        rack_entry, _ = Rack.objects.get_or_create(user=request.user,sku=sku)
                        if difference < 0:
                            rack_entry.quantity += abs(difference)
                            rack_entry.save()
                            InventoryTransaction.objects.create(
                                user=request.user,
                                sku=sku,
                                quantity_change=0,
                                source='bag_edit',
                                notes=f'Reduced in bag {bag.bag_number} (moved {abs(difference)} to rack)'
                            )

                        elif difference > 0:
                            # SKU quantity increased → take from rack
                            if rack_entry.quantity >= difference:
                                rack_entry.quantity -= difference
                                rack_entry.save()
                                InventoryTransaction.objects.create(
                                    user=request.user,
                                    sku=sku,
                                    quantity_change=0,
                                    source='bag_edit',
                                    notes=f'Increased in bag {bag.bag_number} (taken {difference} from rack)'
                                )
                            else:
                                messages.error(request, f"Not enough quantity in rack for SKU {sku.sku_code}")
                                return redirect('edit_bag', pk=bag.pk)

                    else:
                        sku.stock_quantity += difference
                        sku.save()
                        InventoryTransaction.objects.create(
                            user=request.user,
                            sku=sku,
                            quantity_change=difference if not reflect else 0,
                            source='bag_edit',
                            notes=f'Edited in bag {bag.bag_number}' + (' (rack only)' if reflect else '')
                        )

                    existing_item.quantity = quantity
                    existing_item.product_name = sku.product_name
                    existing_item.save()
                else:
                    # ➕ New SKU added to bag
                    BagItem.objects.create(
                        bag=bag,
                        sku=sku,
                        quantity=quantity,
                        product_name=sku.product_name
                    )
                    sku.stock_quantity += quantity
                    sku.save()

                    InventoryTransaction.objects.create(
                        user=request.user,
                        sku=sku,
                        quantity_change=quantity,
                        source="bag_edit",
                        notes=f"Added new SKU to bag {bag.bag_number}"
                    )

            # ❌ Removed SKUs (no longer in formset)
            for removed_item in existing_items.values():
                sku = removed_item.sku
                qty = removed_item.quantity

                if reflect:
                    # Move removed quantity to rack
                    rack_entry, _ = Rack.objects.get_or_create(user=request.user,sku=sku)
                    rack_entry.quantity += qty
                    rack_entry.save()

                    InventoryTransaction.objects.create(
                        user=request.user,
                        sku=sku,
                        quantity_change=0,
                        source="bag_edit",
                        notes=f"Removed from bag {bag.bag_number} → moved {qty} to rack"
                    )
                else:
                    # Deduct from master SKU
                    sku.stock_quantity -= qty
                    sku.save()

                    InventoryTransaction.objects.create(
                        user=request.user,
                        sku=sku,
                        quantity_change=-qty,
                        source="bag_edit",
                        notes=f"Removed from bag {bag.bag_number}"
                    )

                removed_item.delete()

            messages.success(request, "Bag updated and inventory adjusted.")
            return redirect('view_bags')
        else:
            messages.error(request, "Form is invalid.")
    else:
        bag_form = BagForm(initial={'bag_number': bag.bag_number})
        initial_data = [{'sku': item.sku, 'quantity': item.quantity} for item in bag.items.all()]
        formset = BagItemFormSet(initial=initial_data)

    return render(request, 'inventory_manager/edit_bag.html', {
        'bag': bag,
        'bag_form': bag_form,
        'formset': formset,
    })

@login_required
def delete_bag(request, pk):
    bag = get_object_or_404(Bag, pk=pk, user=request.user)

    if request.method == "POST":
        reflect_in_rack = 'reflect_in_rack' in request.POST
        for item in bag.items.all():
            if reflect_in_rack:
                # Move to rack
                rack_entry, _ = Rack.objects.get_or_create(user=request.user,sku=item.sku)
                rack_entry.quantity += item.quantity
                rack_entry.save()

                InventoryTransaction.objects.create(
                    user=request.user,
                    sku=item.sku,
                    quantity_change=0,
                    source="bag_delete",
                    notes=f"Deleted bag {bag.bag_number}, moved {item.quantity} to rack"
                )

            else:
                item.sku.stock_quantity -= item.quantity
                item.sku.save()

                InventoryTransaction.objects.create(
                    user=request.user,
                    sku=item.sku,
                    quantity_change=-item.quantity,
                    source="bag_delete",
                    notes=f"Deleted bag {bag.bag_number}"
                )

        bag.delete()
        messages.success(request, "Bag deleted and inventory adjusted.")
        return redirect("view_bags")
    messages.error(request, "Invalid request method.")
    return redirect('view_bags')

@login_required
def rack_view(request):
    skus = SKU.objects.filter(user=request.user)
    query = request.GET.get("q") or ""
    sort = request.GET.get("sort", "sku__sku_code")  # Default sort
    direction = request.GET.get("dir", "asc")
    per_page = int(request.GET.get("per_page", 25))
    page_number = request.GET.get("page", 1)

    order_by = sort if direction == "asc" else f"-{sort}"

    rack_data = Rack.objects.filter(user=request.user).select_related('sku')

    if query:
        rack_data = rack_data.filter(
            Q(sku__sku_code__icontains=query) |
            Q(sku__product_name__icontains=query)
        )

    rack_data = rack_data.order_by(order_by)

    paginator = Paginator(rack_data, per_page)
    page_obj = paginator.get_page(page_number)

    # 📝 For use in template for sort link retention
    query_params = request.GET.copy()
    if "page" in query_params:
        del query_params["page"]
    query_string = urlencode(query_params)

    if request.method == 'POST':
        for sku in skus:
            key = f"qty_{sku.id}"
            if key in request.POST:
                try:
                    new_qty = int(request.POST[key])
                    rack, created = Rack.objects.get_or_create(user=request.user,sku=sku)
                    diff = new_qty - rack.quantity
                    rack.quantity = new_qty
                    rack.save()

                    sku.stock_quantity += diff
                    sku.save()

                    if diff != 0:
                        InventoryTransaction.objects.create(
                            user=request.user,
                            sku=sku,
                            quantity_change=diff,
                            source="rack_adjust",
                            notes="Manual rack update"
                        )
                except ValueError:
                    pass

        messages.success(request, "Rack quantities updated.")
        return redirect('rack_view')

    return render(request, 'inventory_manager/rack.html', {
        'page_obj': page_obj,
        'query': query,
        'sort': sort,
        'dir': direction,
        'per_page': per_page,
        'query_string': query_string
    })

@login_required
def upload_rack_excel(request):
    if request.method == 'POST' and request.FILES.get('file'):
        mode = request.POST.get("mode", "add")
        file = request.FILES['file']
        file_name = file.name
        short_file_name = file_name[:30]
        unmatched_skus  = []
        try:
            if file_name.endswith(".csv"):
                df = pd.read_csv(file)
            elif file_name.endswith(".xls") or file_name.endswith(".xlsx"):
                df = pd.read_excel(file)
            else:
                messages.error(request, "Unsupported file format.")
                return redirect('rack_view')
        
        except Exception as e:
            messages.error(request, f"Failed to read file: {e}")
            return redirect('rack_view')

        required_cols = ['Master SKU', 'Qty']
        if not all(col in df.columns for col in required_cols):
            messages.error(request, f"Missing columns: {', '.join(required_cols)}")
            return redirect('rack_view')

        for _, row in df.iterrows():
            raw_sku = str(row['Master SKU']).strip()
            qty = int(row['Qty'])

            # Try to find SKU directly
            sku = SKU.objects.filter(user=request.user, sku_code=raw_sku).first()

            # If not found, look for alternate SKU in ChannelListing
            if not sku:
                channel = ChannelListing.objects.filter(user=request.user, channel_sku=raw_sku).first()
                if channel and channel.master_sku:
                    sku = channel.master_sku

            if sku:
                rack, _ = Rack.objects.get_or_create(user=request.user,sku=sku)

                if mode == "add":
                    diff = qty
                    rack.quantity += qty
                elif mode == "subtract":
                    diff = -qty
                    rack.quantity -= qty
                elif mode == "replace":
                    diff = qty - rack.quantity
                    rack.quantity = qty
                else:
                    diff = 0

                rack.save()

                sku.stock_quantity += diff
                sku.save()

                InventoryTransaction.objects.create(
                    user=request.user,
                    sku=sku,
                    quantity_change=diff,
                    source="rack_excel",
                    notes=f"Excel {mode} in rack (via: {raw_sku}/{short_file_name})"
                )
            else:
                unmatched_skus.append({'Master SKU': raw_sku, 'Qty': qty})
                continue

        if unmatched_skus:
                    unmatched_df = pd.DataFrame(unmatched_skus)
                    response = HttpResponse(
                        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                    )
                    response['Content-Disposition'] = 'attachment; filename="Unmatched_SKUs.xlsx"'
                    unmatched_df.to_excel(response, index=False)
                    return response

        messages.success(request, "Rack inventory updated from Excel.")
        return redirect('rack_view')

    return redirect('rack_view')

@login_required
def export_rack_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rack Inventory"

    ws.append(['SKU Code', 'Product Name', 'Rack Quantity'])

    for rack in Rack.objects.filter(user=request.user).select_related('sku'):
        ws.append([rack.sku.sku_code, rack.sku.product_name, rack.quantity])

    for col in range(1, 4):
        ws.column_dimensions[get_column_letter(col)].width = 25

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Rack_Inventory.xlsx'
    wb.save(response)
    return response

@login_required
def ajax_rack_stock(request):
    sku_id = request.GET.get("sku_id")
    try:
        sku = SKU.objects.get(id=sku_id, user=request.user)
        rack = Rack.objects.filter(user=request.user,sku=sku).first()
        qty = rack.quantity if rack else 0
        return JsonResponse({"success": True, "quantity": qty})
    except SKU.DoesNotExist:
        return JsonResponse({"success": False, "message": "SKU not found"})

@login_required
def reset_master_stock(request):
    if request.method == "POST":
        skus = SKU.objects.filter(user=request.user)
        for sku in skus:
            if sku.stock_quantity > 0:
                InventoryTransaction.objects.create(
                    user=request.user,
                    sku=sku,
                    quantity_change=-sku.stock_quantity,
                    source="reset_all",
                    notes="Full reset: stock, bags, and rack"
                )
            sku.stock_quantity = 0
            sku.save()
        # Delete all BagItems and Bags
        Bag.objects.filter(user=request.user).delete()

        #Delete Rack entries
        Rack.objects.filter(user=request.user).delete()
        
        messages.success(request, "✅ All master SKU stock has been reset to 0.")
        return redirect("master_skus")
    
    messages.error(request, "Invalid request method.")
    return redirect("master_skus")