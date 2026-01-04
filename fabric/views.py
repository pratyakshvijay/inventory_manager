from django.shortcuts import render, get_object_or_404, redirect
from .models import Vendor, FabricLot, FabricColorDetail, JobWorkIssue, JobWorkFabricItem, Manufacturer
from .forms import JobWorkFabricItemSizeFormSet, VendorForm, FabricLotForm, FabricColorFormSet, JobWorkIssueForm, JobWorkFabricItemFormSet, ManufacturerForm
import openpyxl
from django.http import HttpResponse, JsonResponse
from django.core.paginator import Paginator
from django.db.models import Q, Sum
from django.contrib import messages
from openpyxl.utils import get_column_letter
from django.core.serializers.json import DjangoJSONEncoder
import json
from decimal import Decimal

def dashboard(request):
    return render(request, 'fabric/dashboard.html')

# Vendor Views
def vendor_list(request):
    vendors = Vendor.objects.all()
    return render(request, 'fabric/vendor_list.html', {'vendors': vendors})

def vendor_create(request):
    form = VendorForm(request.POST or None)
    if form.is_valid():
        form.save()
        return redirect('vendor_list')
    return render(request, 'fabric/vendor_form.html', {'form': form})

def vendor_edit(request, pk):
    vendor = get_object_or_404(Vendor, pk=pk)
    form = VendorForm(request.POST or None, instance=vendor)
    if form.is_valid():
        form.save()
        return redirect('vendor_list')
    return render(request, 'fabric/vendor_form.html', {'form': form})

def vendor_delete(request, pk):
    vendor = get_object_or_404(Vendor, pk=pk)
    vendor.delete()
    return redirect('vendor_list')

# Fabric Lot Views
def fabric_lot_list(request):
    vendor = request.GET.get('vendor')
    invoice_no = request.GET.get('invoice')
    fabric_type = request.GET.get('fabric_type')
    color = request.GET.get('color')
    lot_number = request.GET.get('lot_number')
    per_page = int(request.GET.get('per_page', 10))

    lots = FabricLot.objects.select_related('vendor').prefetch_related('colors').order_by('-lot_number')

    if vendor:
        lots = lots.filter(vendor__name__icontains=vendor)
    if invoice_no:
        lots = lots.filter(invoice_no__icontains=invoice_no)
    if fabric_type:
        lots = lots.filter(fabric_type__icontains=fabric_type)
    if color:
        lots = lots.filter(colors__color__icontains=color)
    if lot_number:
        lots = lots.filter(lot_number__icontains=lot_number)

    lots = lots.distinct()

    # 🔥 If exporting
    if request.GET.get('export') == 'excel':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Fabric Lots"

        headers = ['Lot Number', 'Vendor', 'Invoice No', 'Fabric Type', 'Rate', 'Discount', 'GST', 'Purchase Date', 'Total Meters', 'Total Amount', 'Colors']
        ws.append(headers)

        for lot in lots:
            color_summary = ', '.join(f"{c.color} ({c.meter}m)" for c in lot.colors.all())
            row = [
                lot.lot_number,
                lot.vendor.name,
                lot.invoice_no,
                lot.fabric_type,
                lot.rate,
                lot.discount,
                lot.gst,
                lot.purchase_date.strftime('%Y-%m-%d'),
                lot.total_meters(),
                lot.total_amount(),
                color_summary
            ]
            ws.append(row)

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=FabricLots.xlsx'
        wb.save(response)
        return response

    paginator = Paginator(lots.distinct(), per_page)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'fabric/fabric_lot_list.html', {
        'page_obj': page_obj,
        'lots': page_obj.object_list,  # Required if template uses `lots`
        'vendors': Vendor.objects.all(),
        'selected_vendor': vendor,
        'fabric_type': fabric_type,
        'invoice_no': invoice_no,
        'color': color,
        'lot_number': lot_number,
        'per_page': per_page,
        'request': request,  # Needed by pagination_footer
    })


def fabric_lot_create(request):
    if request.method == 'POST':
        lot_form = FabricLotForm(request.POST)
        color_formset = FabricColorFormSet(request.POST, prefix='form')
        lot_number = request.POST.get('lot_number')
        if FabricLot.objects.filter(lot_number=lot_number).exists():
            messages.error(request, f"Lot Number '{lot_number}' already exists.")
        elif lot_form.is_valid() and color_formset.is_valid():
            lot = lot_form.save()
            color_formset.instance = lot
            color_formset.save()
            return redirect('fabric_lot_list')
    else:
        lot_form = FabricLotForm()
        color_formset = FabricColorFormSet(queryset=FabricColorDetail.objects.none(), prefix='form')

    return render(request, 'fabric/fabric_lot_form.html', {
        'lot_form': lot_form,
        'color_formset': color_formset,
    })


def fabric_lot_edit(request, pk):
    lot = get_object_or_404(FabricLot, pk=pk)
    if request.method == 'POST':
        lot_form = FabricLotForm(request.POST, instance=lot)
        color_formset = FabricColorFormSet(request.POST, instance=lot)
        if lot_form.is_valid() and color_formset.is_valid():
            print("Valid")
            lot_form.save()
            color_formset.save()
            return redirect('fabric_lot_list')
    else:
        lot_form = FabricLotForm(instance=lot)
        color_formset = FabricColorFormSet(instance=lot)

    return render(request, 'fabric/fabric_lot_form.html', {
        'lot_form': lot_form,
        'color_formset': color_formset,
    })

def fabric_lot_delete(request, pk):
    lot = get_object_or_404(FabricLot, pk=pk)
    lot.delete()
    return redirect('fabric_lot_list')

def fabric_lot_export_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fabric Lots"

    # Headers
    ws.append([
        "Lot Number", "Vendor", "Invoice No", "Fabric Type",
        "Rate", "Discount", "GST%", "Purchase Date",
        "Color", "Color Meter", "Total Meters", "Total Amount"
    ])

    for lot in FabricLot.objects.prefetch_related('colors').select_related('vendor'):
        for color in lot.colors.all():
            ws.append([
                lot.lot_number,
                lot.vendor.name,
                lot.invoice_no,
                lot.fabric_type,
                float(lot.rate),
                float(lot.discount),
                float(lot.gst),
                lot.purchase_date.strftime("%Y-%m-%d"),
                color.color,
                float(color.meter),
                float(lot.total_meters()),
                float(lot.total_amount()),
            ])

    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="fabric_lots.xlsx"'
    wb.save(response)
    return response

def job_work_issue_create(request, pk=None):
    instance = JobWorkIssue.objects.get(pk=pk) if pk else None
    edit_mode = bool(pk)
    size_formsets = []  # store size formsets here

    if request.method == 'POST':
        issue_form = JobWorkIssueForm(request.POST, instance=instance)
        item_formset = JobWorkFabricItemFormSet(request.POST, instance=instance, prefix='items')

        if issue_form.is_valid() and item_formset.is_valid():
            issue = issue_form.save()
            item_formset.instance = issue
            items = item_formset.save()

            # Handle size formsets in edit mode
            if edit_mode:
                for i, fabric_item in enumerate(items):
                    size_formset = JobWorkFabricItemSizeFormSet(
                        request.POST,
                        instance=fabric_item,
                        prefix=f'sizes-{i}'
                    )
                    size_formsets.append(size_formset)  # store for re-render if invalid
                    if size_formset.is_valid():
                        size_formset.save()

            messages.success(request, f"Issue {'updated' if edit_mode else 'created'} successfully.")
            return redirect('job_work_issue_list')
        else:
            # Also prepare size formsets if validation fails
            if edit_mode:
                for i, fabric_item in enumerate(instance.items.all()):
                    size_formsets.append(
                        JobWorkFabricItemSizeFormSet(request.POST, instance=fabric_item, prefix=f'sizes-{i}')
                    )
    else:
        issue_form = JobWorkIssueForm(instance=instance)
        item_formset = JobWorkFabricItemFormSet(instance=instance, prefix='items')

        if edit_mode:
            for i, fabric_item in enumerate(instance.items.all()):
                size_formsets.append(
                    JobWorkFabricItemSizeFormSet(instance=fabric_item, prefix=f'sizes-{i}')
                )

    return render(request, 'fabric/job_work_issue_form.html', {
        'issue_form': issue_form,
        'item_formset': item_formset,
        'edit_mode': edit_mode,
        'issue': instance,
        'size_formsets': size_formsets,  # pass to template
    })



def job_work_issue_list(request):
    issues = JobWorkIssue.objects.select_related('manufacturer').prefetch_related('items__fabric_lot', 'items__color_detail').order_by('-issue_date')

    return render(request, 'fabric/job_work_issue_list.html', {
        'issues': issues
    })

def job_work_issue_delete(request, pk):
    issue = get_object_or_404(JobWorkIssue, pk=pk)
    issue.delete()
    messages.success(request, "Job work issue deleted.")
    return redirect('job_work_issue_list')

def get_colors_for_lot(request):
    lot_id = request.GET.get('lot_id')
    if lot_id:
        colors = FabricColorDetail.objects.filter(lot_id=lot_id).values('id', 'color', 'meter')
        return JsonResponse(list(colors), safe=False)
    return JsonResponse([], safe=False)

def get_available_color_details():

    issued_data = (
        JobWorkFabricItem.objects
        .values('color_detail_id')
        .annotate(total_issued=Sum('meter_issued'))
    )
    issued_map = {entry['color_detail_id']: entry['total_issued'] for entry in issued_data}

    result = []
    for color in FabricColorDetail.objects.select_related('lot', 'lot__vendor'):
        total_issued = issued_map.get(color.id, 0)
        remaining = color.meter - Decimal(total_issued)
        if remaining > 0:
            result.append({
                'id': color.id,
                'lot_id': color.lot.id,
                'lot_number': color.lot.lot_number,
                'color': color.color,
                'remaining': remaining,
                'meter': float(remaining),
            })
    return result


def manufacturer_create(request):
    if request.method == 'POST':
        form = ManufacturerForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('manufacturer_list')
    else:
        form = ManufacturerForm()
    return render(request, 'fabric/manufacturer_form.html', {'form': form})


def manufacturer_list(request):
    manufacturers = Manufacturer.objects.all()
    return render(request, 'fabric/manufacturer_list.html', {'manufacturers': manufacturers})

def manufacturer_edit(request, pk):
    manufacturer = get_object_or_404(Manufacturer, pk=pk)
    if request.method == 'POST':
        form = ManufacturerForm(request.POST, instance=manufacturer)
        if form.is_valid():
            form.save()
            messages.success(request, "Manufacturer updated successfully.")
            return redirect('manufacturer_list')
    else:
        form = ManufacturerForm(instance=manufacturer)
    return render(request, 'fabric/manufacturer_form.html', {'form': form, 'edit_mode': True})


def manufacturer_delete(request, pk):
    manufacturer = get_object_or_404(Manufacturer, pk=pk)
    manufacturer.delete()
    messages.success(request, "Manufacturer deleted successfully.")
    return redirect('manufacturer_list')

# Dropdown fetching fabric lot dynamically

def get_available_fabric_lots():
    available_lots = []

    for color in FabricColorDetail.objects.select_related('lot'):
        total_issued = JobWorkFabricItem.objects.filter(color_detail=color).aggregate(
            total=Sum('meter'))['total'] or 0
        remaining = color.meter - total_issued

        if remaining > 0:
            available_lots.append({
                'lot_id': color.lot.id,
                'lot_number': color.lot.lot_number,
                'fabric_type': color.lot.fabric_type,
                'vendor': color.lot.vendor.name,
                'color_id': color.id,
                'color': color.color,
                'remaining_meter': float(remaining),
            })

    return available_lots

def ajax_get_available_lots(request):
    data = get_available_fabric_lots()
    return JsonResponse(data, safe=False)

# End of fetching remiaining fabric lot